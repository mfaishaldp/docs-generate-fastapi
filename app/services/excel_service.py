import math
import numpy as np
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

from app.utils.header_detector import detect_header_row
from app.utils.header_mapper import normalize_headers
from app.utils.table_cleaner import clean_table, trim_table_rows, filter_outlier_rows


def process_excel(
    file_bytes,
    expected_headers,
    header_mapping,
    sheet_name=None,
    sheet_names=None,
    footer_keywords=None,
    min_data_cells=2,
    min_filled_ratio=0.6,
    max_single_cell_chars=60,
    outlier_trim_enabled=True,
    outlier_similarity_threshold=40,
    outlier_max_ratio=0.6,
    outlier_min_columns=2
):

    workbook = load_workbook(
        filename=BytesIO(file_bytes),
        data_only=True
    )

    all_results = []

    if sheet_name:
        target_sheets = [sheet_name]
    elif sheet_names:
        target_sheets = list(sheet_names)
    else:
        target_sheets = list(workbook.sheetnames)

    target_sheets = [
        name
        for name in target_sheets
        if name in workbook.sheetnames
    ]

    # LOOP EACH SHEET
    for sheet_name in target_sheets:

        worksheet = workbook[sheet_name]

        raw_rows = []

        # READ ALL ROWS
        for row in worksheet.iter_rows(values_only=True):

            raw_rows.append(
                list(row)
            )

        # CLEAN TABLE
        cleaned_table = clean_table(
            raw_rows
        )

        if len(cleaned_table) == 0:
            continue

        # DETECT HEADER ROW
        header_index = detect_header_row(
            cleaned_table,
            expected_headers
        )

        if header_index is None:
            continue

        # GET RAW HEADERS
        raw_headers = cleaned_table[
            header_index
        ]

        # NORMALIZE HEADERS
        normalized_headers = normalize_headers(
            raw_headers,
            header_mapping
        )

        # GET DATA ROWS
        rows = cleaned_table[
            header_index + 1:
        ]

        valid_rows = []

        for row in rows:

            # skip fully empty row
            if all(
                str(cell).strip() == ""
                for cell in row
            ):
                continue

            valid_rows.append(row)

        if len(valid_rows) == 0:
            continue

        # ====================================
        # REMOVE INVALID / EMPTY HEADERS
        # ====================================

        filtered_headers = []
        filtered_indexes = []

        for idx, header in enumerate(
            normalized_headers
        ):

            # SKIP NONE HEADER
            if header is None:
                continue

            # SKIP EMPTY HEADER
            if str(header).strip() == "":
                continue

            filtered_headers.append(
                header
            )

            filtered_indexes.append(
                idx
            )

        # SKIP IF NO VALID HEADERS
        if len(filtered_headers) == 0:
            continue

        # ====================================
        # FILTER ROW VALUES
        # ====================================

        filtered_rows = []

        for row in valid_rows:

            cleaned_row = []

            for idx in filtered_indexes:

                if idx < len(row):

                    value = row[idx]

                    # HANDLE NONE
                    if value is None:
                        value = ""

                    cleaned_row.append(
                        str(value).strip()
                    )

                else:

                    cleaned_row.append("")

            filtered_rows.append(
                cleaned_row
            )

        if outlier_trim_enabled:
            filtered_rows = filter_outlier_rows(
                filtered_rows,
                min_columns=outlier_min_columns,
                similarity_threshold=outlier_similarity_threshold,
                max_outlier_ratio=outlier_max_ratio
            )


        if len(filtered_rows) == 0:
            continue

        print(filtered_rows)
        # Remove trailing footer/messy rows
        effective_min_data_cells = min_data_cells
        if len(filtered_headers) == 1:
            effective_min_data_cells = 1

        filtered_rows = trim_table_rows(
            filtered_rows,
            footer_keywords=footer_keywords,
            min_data_cells=effective_min_data_cells,
            expected_columns=len(filtered_headers),
            min_filled_ratio=min_filled_ratio,
            max_single_cell_chars=max_single_cell_chars
        )

        if len(filtered_rows) == 0:
            continue

        # ====================================
        # CREATE DATAFRAME
        # ====================================

        df = pd.DataFrame(
            filtered_rows,
            columns=filtered_headers
        )

        # ====================================
        # REMOVE DUPLICATE COLUMNS
        # ====================================

        df = df.loc[
            :,
            ~df.columns.duplicated()
        ]

        # ====================================
        # REPLACE EMPTY STRING -> None
        # ====================================

        df = df.replace(
            r'^\s*$',
            None,
            regex=True
        )

        # ====================================
        # DROP ROWS WHERE
        # ALL VALUES ARE EMPTY
        # ====================================

        df = df.dropna(
            how="all"
        )

        # ====================================
        # OPTIONAL:
        # REMOVE KEYS WITH NONE VALUE
        # ====================================

        records = df.to_dict(
            orient="records"
        )

        sanitized_records = []

        for record in records:

            sanitized_record = {}

            for key, value in record.items():

                if value is None or value == "":
                    continue

                if isinstance(value, float) and not math.isfinite(value):
                    continue

                if isinstance(value, np.floating):
                    if not math.isfinite(float(value)):
                        continue
                    value = value.item()

                if isinstance(value, np.integer):
                    value = value.item()

                sanitized_record[key] = value

            if len(sanitized_record) == 0:
                continue

            sanitized_records.append(
                sanitized_record
            )

        all_results.extend(
            sanitized_records
        )

    return all_results